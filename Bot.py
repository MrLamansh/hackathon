import asyncio
import pandas as pd
from openpyxl import load_workbook
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart, Command
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
import re
import os
import shutil
from dotenv import load_dotenv
from Generator import ScheduleGenerator
from data_processor import DataProcessor

load_dotenv()

TOKEN = os.getenv("TOKEN")
EXCEL_FILE = os.getenv("EXCEL_FILE", "rasp_prepare_94.xlsx")


def get_user_schedule_file(user_id: int) -> str:
    return f"current_schedule_{user_id}.xlsx"


if not TOKEN:
    raise ValueError("❌ Токен не найден! Создайте файл .env и добавьте TOKEN=your_bot_token")

bot = Bot(token=TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)


class ScheduleStates(StatesGroup):
    choosing_group = State()
    choosing_subgroup = State()
    editing_field = State()
    editing_value = State()
    confirm_edit = State()


class GenerateStates(StatesGroup):
    waiting_for_file = State()
    collecting_exercise_times = State()
    entering_start_time = State()
    confirm_generation = State()


def cleanup_temp_files(data: dict):
    if 'user_file_path' in data and os.path.exists(data['user_file_path']):
        os.remove(data['user_file_path'])
    if 'processed_file' in data and os.path.exists(data['processed_file']):
        os.remove(data['processed_file'])


def load_groups(user_id: int):
    schedule_file = get_user_schedule_file(user_id)
    if not os.path.exists(schedule_file):
        return []

    groups = set()
    try:
        excel_file = pd.ExcelFile(schedule_file)
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(schedule_file, sheet_name=sheet_name)
            if 'Группа' in df.columns:
                for group in df['Группа'].dropna().unique():
                    groups.add(str(group))
        return sorted(list(groups))
    except Exception as e:
        print(f"Ошибка при загрузке групп: {e}")
        return []


def load_subgroups(group_name, user_id: int):
    schedule_file = get_user_schedule_file(user_id)
    if not os.path.exists(schedule_file):
        return []

    subgroups = set()
    try:
        excel_file = pd.ExcelFile(schedule_file)
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(schedule_file, sheet_name=sheet_name)
            if 'Группа' in df.columns and 'Подгруппа' in df.columns:
                # Фильтруем по выбранной группе
                group_df = df[df['Группа'] == group_name]
                for subgroup in group_df['Подгруппа'].dropna().unique():
                    subgroups.add(str(subgroup))
        return sorted(list(subgroups))
    except Exception as e:
        print(f"Ошибка при загрузке подгрупп: {e}")
        return []


def get_schedule_info(group_name, subgroup_name, user_id: int):
    schedule_file = get_user_schedule_file(user_id)
    if not os.path.exists(schedule_file):
        return None

    try:
        excel_file = pd.ExcelFile(schedule_file)
        stages_info = []

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(schedule_file, sheet_name=sheet_name)

            if 'Группа' in df.columns and 'Подгруппа' in df.columns:
                matches = df[(df['Группа'] == group_name) & (df['Подгруппа'] == subgroup_name)]

                for _, row in matches.iterrows():
                    stages_info.append({
                        'court': sheet_name,
                        'time': row.get('Время', '—'),
                        'stage': row.get('Этап', '—'),
                        'participants': row.get('Участников', '—'),
                        'poomse': row.get('Пхумсе', '—')
                    })

        if not stages_info:
            return None

        stages_info.sort(key=lambda x: x['time'] if x['time'] != '—' else '99:99')

        first_stage = stages_info[0]

        all_poomse = []
        stage_details = []
        for stage in stages_info:
            poomse = str(stage['poomse'])
            if poomse and poomse != '—' and poomse not in all_poomse:
                all_poomse.append(poomse)
            stage_details.append(f"{stage['stage']} ({stage['time']}, {stage['court']})")

        return {
            "kort": first_stage['court'],
            "start_time": first_stage['time'],
            "participants": str(first_stage['participants']),
            "poomse": ", ".join(all_poomse) if all_poomse else "—",
            "stages": " → ".join(stage_details)
        }

    except Exception as e:
        print(f"Ошибка при получении информации о расписании: {e}")
        return None


def update_excel_cell(sheet_name, row_idx, col_idx, value):
    wb = load_workbook(EXCEL_FILE)
    if sheet_name not in wb.sheetnames:
        return False
    ws = wb[sheet_name]
    # openpyxl индексирует с 1
    ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)
    wb.save(EXCEL_FILE)
    return True


# === Обработчики просмотра/редактирования ===
@dp.message(CommandStart())
async def start(message: types.Message, state: FSMContext):
    buttons = [
        [KeyboardButton(text="📅 Просмотреть расписание")],
        [KeyboardButton(text="🔧 Сгенерировать новое расписание")],
        [KeyboardButton(text="❌ Отмена")]
    ]
    keyboard = ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)
    await message.answer("🏆 Выберите действие:", reply_markup=keyboard)
    await state.clear()


@dp.message(F.text == "📅 Просмотреть расписание")
async def view_schedule(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    schedule_file = get_user_schedule_file(user_id)

    # Проверяем наличие сгенерированного расписания
    if not os.path.exists(schedule_file):
        await message.answer(
            "❌ Расписание ещё не было сгенерировано.\n\n"
            "Пожалуйста, сначала выберите '🔧 Сгенерировать новое расписание' "
            "и создайте расписание, загрузив свой файл."
        )
        return

    groups = load_groups(user_id)
    if not groups:
        await message.answer("❌ Не удалось загрузить группы из расписания.")
        return

    buttons = [[KeyboardButton(text=g)] for g in groups]
    buttons.append([KeyboardButton(text="🔙 Назад")])
    keyboard = ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)
    await message.answer("🏆 Выберите группу:", reply_markup=keyboard)
    await state.set_state(ScheduleStates.choosing_group)


@dp.message(ScheduleStates.choosing_group)
async def choose_group(message: types.Message, state: FSMContext):
    if message.text in ["❌ Отмена", "🔙 Назад"]:
        await start(message, state)
        return

    user_id = message.from_user.id
    groups = load_groups(user_id)
    if message.text not in groups:
        await message.answer("❌ Такой группы нет. Выберите из списка.")
        return

    await state.update_data(selected_group=message.text)
    subgroups = load_subgroups(message.text, user_id)
    if not subgroups:
        await message.answer("❌ Подгруппы не найдены.")
        return

    buttons = [[KeyboardButton(text=s)] for s in subgroups]
    buttons.append([KeyboardButton(text="🔙 Назад к группам")])
    keyboard = ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)
    await message.answer(f"Группа: *{message.text}*\nВыберите подгруппу:", reply_markup=keyboard, parse_mode="Markdown")
    await state.set_state(ScheduleStates.choosing_subgroup)


@dp.message(ScheduleStates.choosing_subgroup)
async def choose_subgroup(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await start(message, state)
        return

    if message.text == "🔙 Назад к группам":
        await view_schedule(message, state)
        return

    if message.text == "🔙 Назад":
        await start(message, state)
        return

    user_id = message.from_user.id
    data = await state.get_data()
    group = data.get("selected_group")
    info = get_schedule_info(group, message.text, user_id)
    if not info:
        await message.answer("❌ Подгруппа не найдена.")
        return

    await state.update_data(current_info=info, selected_subgroup=message.text)

    text = (
        f"📋 *Расписание выступления*\n\n"
        f"🏷 Группа: `{group}`\n"
        f"🔖 Подгруппа: `{message.text}`\n"
        f"🏟 Корт: `{info['kort']}`\n"
        f"⏰ Время начала: `{info['start_time']}`\n"
        f"👥 Участников: `{info['participants']}`\n"
        f"🥋 Пхумсе: `{info['poomse']}`\n\n"
        f"📍 Этапы: `{info['stages']}`"
    )

    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🔙 Назад")]
        ],
        resize_keyboard=True
    )
    await message.answer(text, reply_markup=keyboard, parse_mode="Markdown")


@dp.message(F.text == "🔙 Назад")
async def back_handler(message: types.Message, state: FSMContext):
    # Всегда возвращаемся на главный экран
    await start(message, state)


@dp.message(F.text == "✏️ Редактировать")
async def edit_schedule(message: types.Message, state: FSMContext):
    fields = ["⏰ Время начала", "👥 Участников", "🥋 Пхумсе", "🏟 Корт"]
    buttons = [[KeyboardButton(text=f)] for f in fields]
    buttons.append([KeyboardButton(text="❌ Отмена")])
    keyboard = ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)
    await message.answer("Выберите поле для редактирования:", reply_markup=keyboard)
    await state.set_state(ScheduleStates.editing_field)


@dp.message(ScheduleStates.editing_field)
async def choose_edit_field(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await start(message, state)
        return

    field_map = {
        "⏰ Время начала": "start_time",
        "👥 Участников": "participants",
        "🥋 Пхумсе": "poomse",
        "🏟 Корт": "kort"
    }
    internal_field = field_map.get(message.text)
    if not internal_field:
        await message.answer("❌ Неверный выбор.")
        return

    await state.update_data(editing_field=internal_field)
    prompts = {
        "start_time": "Введите новое время начала (формат ЧЧ:ММ, например 10:30):",
        "participants": "Введите новое количество участников (целое число):",
        "poomse": "Введите пхумсе через запятую (например: тхэгук иль джан, кибон иль джан):",
        "kort": "Выберите корт:\n1 — Корт 1\n2 — Корт 2\n3 — Корт 3"
    }
    await message.answer(prompts[internal_field])
    await state.set_state(ScheduleStates.editing_value)


@dp.message(ScheduleStates.editing_value)
async def input_new_value(message: types.Message, state: FSMContext):
    data = await state.get_data()
    field = data["editing_field"]
    value = message.text.strip()

    # Валидация
    if field == "start_time":
        if not re.match(r"^\d{1,2}:\d{2}$", value):
            await message.answer("❌ Неверный формат. Используйте ЧЧ:ММ")
            return
    elif field == "participants":
        if not value.isdigit() or int(value) <= 0:
            await message.answer("❌ Введите положительное целое число.")
            return
    elif field == "kort":
        if value not in ["1", "2", "3"]:
            await message.answer("❌ Введите 1, 2 или 3.")
            return

    await state.update_data(new_value=value)

    # Подтверждение
    display_names = {"start_time": "время начала", "participants": "количество участников", "poomse": "пхумсе",
                     "kort": "корт"}
    confirm_text = f"Изменить *{display_names[field]}* на:\n`{value}`?"
    await message.answer(confirm_text, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да", callback_data="confirm_edit")],
        [InlineKeyboardButton(text="❌ Нет", callback_data="cancel_edit")]
    ]))
    await state.set_state(ScheduleStates.confirm_edit)


@dp.callback_query(F.data == "confirm_edit")
async def confirm_edit(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    field = data["editing_field"]
    new_value = data["new_value"]
    row_idx = data["current_info"]["row_index"]

    # Преобразуем значение для Excel
    if field == "kort":
        letter_map = {"1": "k", "2": "u", "3": "d"}
        excel_value = letter_map[new_value]
        col_idx = 0
    elif field == "start_time":
        excel_value = new_value
        col_idx = 7
    elif field == "participants":
        excel_value = int(new_value)
        col_idx = 5
    elif field == "poomse":
        poomse_list = [p.strip() for p in new_value.split(",") if p.strip()]

        for i, p in enumerate(poomse_list[:3]):
            update_excel_cell("prep", row_idx, 8 + i, p)
        for i in range(len(poomse_list), 3):
            update_excel_cell("prep", row_idx, 8 + i, "")
        await callback.message.edit_text("✅ Расписание успешно обновлено!")
        await start(callback.message, state)
        return

    success = update_excel_cell("prep", row_idx, col_idx, excel_value)
    if success:
        await callback.message.edit_text("✅ Расписание успешно обновлено!")
    else:
        await callback.message.edit_text("❌ Ошибка при сохранении.")

    await start(callback.message, state)


@dp.callback_query(F.data == "cancel_edit")
async def cancel_edit(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("❌ Изменения отменены.")
    await start(callback.message, state)


# === Обработчики генерации расписания ===
@dp.message(F.text == "🔧 Сгенерировать новое расписание")
async def start_generation(message: types.Message, state: FSMContext):
    await message.answer("🔄 Начинаю процесс генерации расписания...")

    template_file = FSInputFile("template.xlsx")
    await message.answer_document(
        template_file,
        caption="📄 Это шаблон файла для заполнения. Скачайте его и заполните своими данными."
    )

    await message.answer(
        "⚠️ *ВАЖНО!*\n\n"
        "Обратите внимание, что слот \"наименование группы\" должен совпадать на обоих листах.\n\n"
        "📋 *Структура файла:*\n\n"
        "*Лист 1 (Группы):*\n"
        "• Столбец 1: Наименование группы\n"
        "• Столбец 2: Подгруппа\n"
        "• Столбец 3: Количество участников\n\n"
        "*Лист 2 (Упражнения):*\n"
        "• Столбец 1: Наименование группы (должно совпадать с Листом 1!)\n"
        "• Столбец 2: Упражнение для отбора (если нужен)\n"
        "• Столбец 3: Упражнение для полуфинала (если нужен)\n"
        "• Столбец 4: Упражнение для финала (обязательно)\n\n"
        "📎 Пожалуйста, отправьте заполненный файл Excel.",
        parse_mode="Markdown"
    )

    await state.set_state(GenerateStates.waiting_for_file)


@dp.message(GenerateStates.waiting_for_file, F.document)
async def process_uploaded_file(message: types.Message, state: FSMContext):
    document = message.document

    if not (document.file_name.endswith('.xlsx') or document.file_name.endswith('.xls')):
        await message.answer("❌ Пожалуйста, отправьте файл Excel (.xlsx или .xls)")
        return

    await message.answer("⏳ Обрабатываю файл...")

    try:
        file_path = f"user_upload_{message.from_user.id}.xlsx"
        await message.bot.download(document, destination=file_path)

        # Обрабатываем файл через DataProcessor
        processor = DataProcessor(file_path)
        success, processed_file, exercises = processor.process(f"processed_{message.from_user.id}.xlsx")

        if not success or not exercises:
            await message.answer("❌ Ошибка при обработке файла. Проверьте структуру данных.")
            # Удаляем временные файлы
            if os.path.exists(file_path):
                os.remove(file_path)
            await start(message, state)
            return

        # Сохраняем данные в состояние
        await state.update_data(
            user_file_path=file_path,
            processed_file=processed_file,
            exercises=exercises,
            exercise_times={},
            current_exercise_index=0
        )

        await message.answer(
            f"✅ Файл успешно обработан!\n"
            f"Найдено {len(exercises)} уникальных упражнений.\n\n"
            f"Теперь мне нужно узнать время выполнения каждого упражнения."
        )

        await ask_exercise_time(message, state)

    except Exception as e:
        await message.answer(f"❌ Ошибка при обработке файла: {str(e)}")
        # Удаляем временные файлы
        if os.path.exists(file_path):
            os.remove(file_path)
        await start(message, state)


async def ask_exercise_time(message: types.Message, state: FSMContext):
    data = await state.get_data()
    exercises = data['exercises']
    current_index = data['current_exercise_index']

    if current_index >= len(exercises):
        await ask_start_time(message, state)
        return

    current_exercise = exercises[current_index]
    await message.answer(
        f"⏱ Упражнение: *{current_exercise}*\n\n"
        f"Введите время выполнения в минутах (например: 1.5 или 2):\n"
        f"Прогресс: {current_index + 1}/{len(exercises)}",
        parse_mode="Markdown"
    )
    await state.set_state(GenerateStates.collecting_exercise_times)


@dp.message(GenerateStates.collecting_exercise_times)
async def collect_exercise_time(message: types.Message, state: FSMContext):
    # Валидация времени
    try:
        time_value = float(message.text.strip().replace(',', '.'))
        if time_value <= 0:
            await message.answer("❌ Время должно быть положительным числом. Попробуйте еще раз.")
            return
    except ValueError:
        await message.answer("❌ Неверный формат. Введите число (например: 1.5 или 2)")
        return

    # Сохраняем время
    data = await state.get_data()
    exercises = data['exercises']
    current_index = data['current_exercise_index']
    exercise_times = data['exercise_times']

    current_exercise = exercises[current_index]
    exercise_times[current_exercise] = time_value

    # Переходим к следующему упражнению
    await state.update_data(
        exercise_times=exercise_times,
        current_exercise_index=current_index + 1
    )

    await ask_exercise_time(message, state)


async def ask_start_time(message: types.Message, state: FSMContext):
    """Спрашивает время начала соревнований"""
    await message.answer(
        "✅ Все упражнения настроены!\n\n"
        "⏰ Теперь введите время начала соревнований в формате ЧЧ:ММ (например: 08:30):"
    )
    await state.set_state(GenerateStates.entering_start_time)


@dp.message(GenerateStates.entering_start_time)
async def collect_start_time(message: types.Message, state: FSMContext):
    # Валидация времени
    if not re.match(r"^\d{1,2}:\d{2}$", message.text.strip()):
        await message.answer("❌ Неверный формат. Используйте ЧЧ:ММ (например: 08:30)")
        return

    start_time = message.text.strip()
    await state.update_data(start_time=start_time)

    # Показываем подтверждение
    data = await state.get_data()
    exercise_times = data['exercise_times']

    summary = "📋 *Сводка параметров:*\n\n"
    summary += f"⏰ Время начала: `{start_time}`\n\n"
    summary += "*Время выполнения упражнений:*\n"
    for ex, time in exercise_times.items():
        summary += f"• {ex}: {time} мин\n"

    summary += "\n🔧 Сгенерировать расписание?"

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да, сгенерировать", callback_data="generate_schedule")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_generation")]
    ])

    await message.answer(summary, parse_mode="Markdown", reply_markup=keyboard)
    await state.set_state(GenerateStates.confirm_generation)


@dp.callback_query(F.data == "generate_schedule")
async def generate_schedule(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("⏳ Генерирую расписание, пожалуйста подождите...")

    try:
        data = await state.get_data()
        processed_file = data['processed_file']
        exercise_times = data['exercise_times']
        start_time = data['start_time']

        # Создаём генератор с обработанным файлом
        generator = ScheduleGenerator(processed_file)

        # Устанавливаем время упражнений
        generator.set_exercise_times(exercise_times)

        # Генерируем расписание
        schedule = generator.generate_schedule(start_time)

        if not schedule:
            await callback.message.answer("❌ Не удалось сгенерировать расписание. Проверьте данные в Excel.")
            # Очищаем временные файлы
            cleanup_temp_files(data)
            await start(callback.message, state)
            return

        # Сохраняем в файл
        output_file = generator.save_schedule_to_excel(schedule, f"schedule_{callback.from_user.id}.xlsx")

        # Копируем файл для функции просмотра (персональный файл пользователя)
        user_schedule_file = get_user_schedule_file(callback.from_user.id)
        shutil.copy(output_file, user_schedule_file)

        # Формируем сводку
        total_slots = len(schedule)
        courts = {1: 0, 2: 0, 3: 0}
        for slot in schedule:
            courts[slot.court] += 1

        end_time = max(slot.end_time for slot in schedule)

        summary = (
            f"✅ *Расписание успешно сгенерировано!*\n\n"
            f"📊 Статистика:\n"
            f"• Всего выступлений: {total_slots}\n"
            f"• Корт 1: {courts[1]} выступлений\n"
            f"• Корт 2: {courts[2]} выступлений\n"
            f"• Корт 3: {courts[3]} выступлений\n"
            f"• Начало: {start_time}\n"
            f"• Окончание: {end_time.strftime('%H:%M')}\n"
        )

        await callback.message.answer(summary, parse_mode="Markdown")

        # Отправляем расписание для каждого корта
        for court_num in [1, 2, 3]:
            court_schedule_text = generator.format_schedule_as_text(schedule, court_num)

            # Разбиваем на части, если текст слишком длинный (лимит Telegram - 4096 символов)
            max_length = 4000  # Оставляем запас
            if len(court_schedule_text) <= max_length:
                await callback.message.answer(court_schedule_text, parse_mode="Markdown")
            else:
                # Разбиваем по блокам времени
                parts = court_schedule_text.split('\n\n')
                current_part = f"*КОРТ {court_num}* (часть 1)\n" + "━" * 50 + "\n\n"
                part_num = 1

                for block in parts[1:]:  # Пропускаем заголовок
                    if len(current_part) + len(block) + 2 > max_length:
                        # Отправляем текущую часть
                        await callback.message.answer(current_part, parse_mode="Markdown")
                        await asyncio.sleep(0.3)
                        part_num += 1
                        current_part = f"*КОРТ {court_num}* (часть {part_num})\n" + "━" * 50 + "\n\n"

                    current_part += block + "\n\n"

                # Отправляем последнюю часть
                if current_part.strip():
                    await callback.message.answer(current_part, parse_mode="Markdown")

            await asyncio.sleep(0.5)  # Небольшая задержка между кортами

        # Отправляем файл
        file = FSInputFile(output_file)
        await callback.bot.send_document(callback.message.chat.id, file, caption="📄 Полное расписание в Excel")

        # Удаляем выходной файл после отправки
        if os.path.exists(output_file):
            os.remove(output_file)

        # Очищаем временные файлы
        cleanup_temp_files(data)

    except Exception as e:
        await callback.message.answer(f"❌ Ошибка при генерации: {str(e)}")
        # Очищаем временные файлы в случае ошибки
        data = await state.get_data()
        cleanup_temp_files(data)

    await start(callback.message, state)


@dp.callback_query(F.data == "cancel_generation")
async def cancel_generation(callback: types.CallbackQuery, state: FSMContext):
    # Очищаем временные файлы
    data = await state.get_data()
    cleanup_temp_files(data)

    await callback.message.edit_text("❌ Генерация отменена.")
    await start(callback.message, state)


# === Запуск ===
async def main():
    print("✅ Бот запущен!")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
